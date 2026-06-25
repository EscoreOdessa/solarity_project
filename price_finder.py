#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EScore Energy — Автоматизатор пошуку цін для СЕС
=================================================
Алгоритм:
  1. Зчитує вкладку «Кошторис_Наявність обладнання» з Google Sheets
  2. Перевіряє прайс ETI (знижка 25%) → бере ціну звідти
  3. Якщо ETI не знайшов — шукає на 3 Ukrainian сайтах через браузер,
     усереднює і записує в таблицю з посиланням на джерело

══════════════════════════════════════════════════════════════
 ОДНОРАЗОВЕ НАЛАШТУВАННЯ (виконати один раз):
   pip install gspread google-auth google-auth-oauthlib \\
               google-api-python-client openpyxl rapidfuzz \\
               colorama playwright playwright-stealth
   playwright install chromium

 НАЛАШТУВАННЯ Google API:
   1. https://console.cloud.google.com → новий проект
   2. Увімкніть «Google Sheets API» та «Google Drive API»
   3. Credentials → OAuth 2.0 Client IDs → Desktop App → Download JSON
   4. Перейменуйте завантажений файл у credentials.json
   5. Покладіть credentials.json поруч із цим скриптом
══════════════════════════════════════════════════════════════
"""

# ╔═══════════════════════════════════════════════════════════╗
# ║   ↓↓↓  ВСТАВТЕ ПОСИЛАННЯ НА GOOGLE SHEETS СЮДИ  ↓↓↓    ║
SHEET_URL = "https://docs.google.com/spreadsheets/d/1CM9f2xcyIr9l_64hLZ6BiugC1GG71AwkWlepUYcux_0/edit?usp=sharing"
# ║   ↑↑↑  БІЛЬШЕ НІЧОГО НЕ ЗМІНЮВАТИ              ↑↑↑    ║
# ╚═══════════════════════════════════════════════════════════╝

# ─── Налаштування (змінювати рідко) ────────────────────────
CREDENTIALS_FILE = "credentials.json"  # OAuth ключ від Google Cloud
TOKEN_FILE       = "token.json"        # створюється автоматично після 1-го входу
SERPER_KEY_FILE  = "serper_api_key.txt"  # файл з ключем Serper (поруч зі скриптом)
SERVICE_ACCOUNT_FILE = "service_account.json"  # ключ сервісного акаунта (пріоритетний вхід)
ETI_FILE_SEARCH  = "eti_04.03.2026"   # рядок пошуку ETI-файлу на Google Drive
ETI_TAB_NAME     = "Price_04.03.2026" # назва вкладки у файлі ETI
ETI_DISCOUNT     = 0.25               # знижка від ETI = 25%
SPEC_TAB_NAME    = "Довідник обладнання та цін"

# Категорії (значення колонки «Тип товару»), для яких НЕ перевіряємо прайс ETI.
# Уся кабельно-провідникова продукція та кабельні аксесуари (короби, лотки,
# канали, конектори) одразу йдуть у веб-пошук: потрібних брендів (KOPOS, EKD,
# E.NEXT, DKC) у прайсі ETI немає, тож fuzzy-match давав хибні збіги
# (усі 15 позицій приліплювались до «Кабельний ввід KVR-OC» = 481.95 грн).
ETI_SKIP_TYPE_KEYWORDS = ["кабель", "короб", "лоток", "конектор", "коннектор"]

MAX_SITES        = 5                  # перевіряємо перші 20 сайтів (за інструкцією)
REQUEST_DELAY    = 4.0                # пауза між запитами (сек) — більше = рідше капча
HEADLESS_BROWSER = False               # False = видно браузер під час пошуку

# Сайти що ІГНОРУЄМО при пошуку цін
SKIP_PRICE_SITES = ["alibaba.com", "aliexpress.com", "eti.ua", "olx.ua"]
# Позиції що ІГНОРУЄМО (перевіряємо за частиною назви)
SKIP_SECTIONS = [
     #"Активний споживач",
     #"Заземлення",
     #"Блискавкозахист",
     "Витратні Витратні матеріали",
     "БМС",
    # "Основне обладнання",
    # "Додаткові витрати",
    # "Кріплення"
]

# Розділи що ІГНОРУЄМО (ціни фіксують інженери або це роботи)
SKIP_ITEMS = [

#     "Основне обладнання",
#     "Сонячні панелі",
#     "Інвертор",
#     "Криплення",
#     "Облік Енергосервіс",
    # "Активний споживач",
    # "Заземлення",
    # "Блискавкозахист",
    "Витратні Витратні матеріали"
    "БМС",
    "БМС BMS Deye",
#     "Додаткові витрати",
#     "витратні матеріали",
#     "Доставка обладнання від постачальника", 
#     "Сонячний кабель",
#     "Кріплення Баластна система без підйому кута",
#     "KBE Solar DB+ 6,00 Q verz EN 50618, IEC 62930, 2 PFG 1169/10.19 black (SW)"
]

# Сайти для пошуку (порядок = пріоритет)
SEARCH_SITES = [
    {"name": "Isolar",       "url": "https://isolar.com.ua/search/?q={q}"},
    {"name": "Eleksun",      "url": "https://eleksun.com.ua/search/?q={q}"},
    {"name": "Ecshop",       "url": "https://ecshop.com.ua/ua/search/?text={q}"},
    {"name": "001",          "url": "https://001.com.ua/ua/search?query={q}"},
    {"name": "Electrocontrol","url": "https://electrocontrol.com.ua/ua/search?search_term={q}"},
    {"name": "Electrovoz",   "url": "https://electrovoz.com.ua/ua/search?search_term={q}"},
    {"name": "E-server",     "url": "https://e-server.com.ua/ua/search?search_term={q}"},
    {"name": "Sunservis",    "url": "https://sunservis.com.ua/ua/search?search_term={q}"},
    {"name": "Abatareykin",  "url": "https://abatareykin.kiev.ua/ua/search?search_term={q}"},
    {"name": "Agart",        "url": "https://agart.ua/ua/search?search_term={q}"},
    {"name": "Pulsepro",     "url": "https://pulsepro.com.ua/ua/search?search_term={q}"},
    {"name": "Acko",         "url": "https://agart.ua/ua/search?search_term={q}"},
    {"name": "Energomotive", "url": "https://energomotive.com.ua/ua/search?search_term={q}"},
    #{"name": "Amperok",    "url": "https://amperok.com.ua/ua/search?search_term={q}"},
    {"name": "Cms",          "url": "https://cms.ua/ua/search?search_term={q}"},
]
# ───────────────────────────────────────────────────────────

import asyncio
import io
import json
import os
import re
import sys
import time
import urllib.request
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote_plus, urlparse

# ── Перевірка залежностей ──────────────────────────────────
_missing = []
try:
    import gspread
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
except ImportError:
    _missing.append("gspread google-auth google-auth-oauthlib google-api-python-client")

try:
    import openpyxl
except ImportError:
    _missing.append("openpyxl")

try:
    from rapidfuzz import fuzz
    from rapidfuzz import process as rfuzz_process
except ImportError:
    _missing.append("rapidfuzz")

try:
    from colorama import Fore, Style, init as colorama_init
    colorama_init(autoreset=True)
except ImportError:
    _missing.append("colorama")
    # Fallback без кольорів
    class Fore:
        RED = YELLOW = GREEN = CYAN = WHITE = RESET = ""
    class Style:
        BRIGHT = RESET_ALL = ""

if _missing:
    print("❌ Відсутні бібліотеки. Встановіть:")
    for pkg in _missing:
        print(f"   pip install {pkg}")
    print("\nТакож запустіть:")
    print("   pip install playwright && playwright install chromium")
    sys.exit(1)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]


# ══════════════════════════════════════════════════════════
# 1. GOOGLE АВТОРИЗАЦІЯ
# ══════════════════════════════════════════════════════════

def get_credentials():
    """
    Авторизація Google. Порядок пріоритету:
      1) Сервісний акаунт з env-змінної GOOGLE_SERVICE_ACCOUNT_JSON (для хмари/розкладу)
      2) Сервісний акаунт з файлу service_account.json (локально, будь-який користувач)
      3) Фолбек: інтерактивний OAuth (credentials.json + браузер) — лише для розробника

    Сервісний акаунт працює без браузера й логіна — головне, щоб таблиця та файл ETI
    були розшарені на email робота (solar-search@…iam.gserviceaccount.com).
    """
    # ── 1 + 2: сервісний акаунт (без інтерактивного входу) ──────────
    sa_env  = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    sa_path = Path(SERVICE_ACCOUNT_FILE)
    if sa_env or sa_path.exists():
        from google.oauth2 import service_account
        if sa_env:
            creds = service_account.Credentials.from_service_account_info(
                json.loads(sa_env), scopes=SCOPES)
            print(f"{Fore.GREEN}✅ Авторизація через сервісний акаунт (секрет оточення)")
        else:
            creds = service_account.Credentials.from_service_account_file(
                str(sa_path), scopes=SCOPES)
            print(f"{Fore.GREEN}✅ Авторизація через сервісний акаунт: {sa_path.name}")
        return creds

    # ── 3: фолбек — інтерактивний OAuth (локальна розробка) ─────────
    creds_path = Path(CREDENTIALS_FILE)
    token_path = Path(TOKEN_FILE)

    if not creds_path.exists():
        print(f"\n{Fore.RED}❌ Немає ні {SERVICE_ACCOUNT_FILE}, ні {CREDENTIALS_FILE}!")
        print(f"   Поклади {SERVICE_ACCOUNT_FILE} поруч зі скриптом (рекомендований спосіб)")
        print("   або налаштуй OAuth credentials.json.\n")
        sys.exit(1)

    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("🔄 Оновлюю токен...")
            creds.refresh(Request())
        else:
            print("🌐 Відкриваю браузер для авторизації Google...")
            flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json())

    print(f"{Fore.GREEN}✅ Авторизовано в Google (OAuth)")
    return creds


# ══════════════════════════════════════════════════════════
# 2. ETI ПРАЙС-ЛИСТ
# ══════════════════════════════════════════════════════════

class ETIPriceList:
    """Завантажує прайс ETI з Google Drive і шукає позиції з fuzzy-match."""

    def __init__(self, creds: Credentials):
        self.creds = creds
        self._items: List[Dict] = []  # [{code, name, price_uah}]
        self.loaded = False

    # ── Завантаження файлу ─────────────────────────────────
    def load(self) -> bool:
        try:
            drive = build("drive", "v3", credentials=self.creds)

            # Знаходимо файл по частині назви
            q = f"name contains '{ETI_FILE_SEARCH}' and trashed=false"
            results = drive.files().list(
                q=q, orderBy="modifiedTime desc",
                pageSize=5, fields="files(id,name,modifiedTime)"
            ).execute()

            files = results.get("files", [])
            if not files:
                print(f"{Fore.YELLOW}⚠  ETI файл '{ETI_FILE_SEARCH}' не знайдено на Drive")
                return False

            file_id = files[0]["id"]
            fname   = files[0]["name"]
            print(f"📋 ETI прайс: {fname}")

            # Завантажуємо в пам'ять
            req = drive.files().get_media(fileId=file_id)
            buf = io.BytesIO()
            dl = MediaIoBaseDownload(buf, req)
            done = False
            while not done:
                _, done = dl.next_chunk()

            buf.seek(0)
            return self._parse(buf)

        except Exception as ex:
            print(f"{Fore.YELLOW}⚠  Помилка завантаження ETI: {ex}")
            return False

    # ── Парсинг xlsx ───────────────────────────────────────
    def _parse(self, buf: io.BytesIO) -> bool:
        try:
            wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)

            # Знайти потрібну вкладку
            ws = None
            for sname in wb.sheetnames:
                if ETI_TAB_NAME.lower() in sname.lower():
                    ws = wb[sname]
                    break
            ws = ws or wb.active

            # Знайти рядок заголовків
            col_code = col_name = col_price = None
            header_row_idx = None

            for ridx, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = [str(c).lower() if c else "" for c in row]
                row_text = " ".join(cells)

                if ("найменування" in row_text or "назва" in row_text) and "ціна" in row_text:
                    header_row_idx = ridx
                    for cidx, cv in enumerate(cells):
                        if not cv:
                            continue
                        if col_code  is None and "код"              in cv:
                            col_code = cidx
                        if col_name  is None and ("найменування" in cv or "назва" in cv):
                            col_name = cidx
                        if col_price is None and "грн" in cv and "ціна" in cv:
                            col_price = cidx
                    break

            if not header_row_idx or col_name is None or col_price is None:
                # Fallback: ищем столбцы по другой логике
                # Часто в ETI файлах первые строки - шапка компании
                # попробуем найти строку с "Ціна в грн"
                for ridx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    cells = [str(c).lower() if c else "" for c in row]
                    row_text = " ".join(cells)
                    if "ціна в грн" in row_text or "price" in row_text:
                        header_row_idx = ridx
                        for cidx, cv in enumerate(cells):
                            if not cv: continue
                            if col_code is None and "код" in cv: col_code = cidx
                            if col_name is None and ("найменування" in cv or "повне" in cv): col_name = cidx
                            if col_price is None and "грн" in cv: col_price = cidx
                        break

            if not header_row_idx or col_name is None or col_price is None:
                print(f"{Fore.YELLOW}⚠  Не знайдено заголовки у ETI файлі (вкладка: {ws.title})")
                print(f"   Очікувані заголовки: 'Повне найменування', 'Ціна в грн. з ПДВ'")
                return False

            # Читаємо позиції
            count = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                name  = row[col_name]  if col_name  < len(row) else None
                price = row[col_price] if col_price < len(row) else None
                code  = row[col_code]  if col_code is not None and col_code < len(row) else None

                if not name or not price:
                    continue

                name_str = str(name).strip()
                if not name_str or name_str.lower() in ("повне найменування", "назва"):
                    continue

                try:
                    price_val = float(
                        str(price).replace(",", ".").replace(" ", "").replace("\xa0", "")
                    )
                    if price_val > 0:
                        self._items.append({
                            "code":  str(code or "").strip(),
                            "name":  name_str,
                            "price": price_val,
                        })
                        count += 1
                except (ValueError, TypeError):
                    pass

            print(f"{Fore.GREEN}✅ ETI прайс завантажено: {count} позицій")
            self.loaded = True
            return True

        except Exception as ex:
            print(f"{Fore.YELLOW}⚠  Помилка парсингу ETI: {ex}")
            return False

    # ── Пошук ціни ─────────────────────────────────────────
    def find(self, item_name: str, threshold: int = 62
             ) -> Optional[Tuple[float, str, str]]:
        """
        Шукає позицію у прайсі.
        Повертає (ціна_зі_знижкою, назва_в_прайсі, примітка) або None.
        """
        if not self.loaded or not self._items:
            return None

        # 1. Пошук за артикулом (число у дужках 7+ цифр)
        code_m = re.search(r'\((\d{6,10})\)', item_name)
        if code_m:
            target_code = code_m.group(1)
            for it in self._items:
                if it["code"] == target_code:
                    return self._result(it, reason=f"артикул {target_code}")

        # 2. Fuzzy match за назвою
        names = [it["name"] for it in self._items]
        match = rfuzz_process.extractOne(
            item_name, names,
            scorer=fuzz.partial_ratio,
            score_cutoff=threshold,
        )
        if match:
            matched_name, score, idx = match
            return self._result(self._items[idx], reason=f"fuzzy {score:.0f}%")

        return None

    def _result(self, item: Dict, reason: str) -> Tuple[float, str, str]:
        discounted = round(item["price"] * (1 - ETI_DISCOUNT), 2)
        note = (
            f"ETI прайс {item['price']:.2f} грн "
            f"× {1 - ETI_DISCOUNT:.2f} (знижка {int(ETI_DISCOUNT * 100)}%) "
            f"| {reason} | {item['name'][:60]}"
        )
        return discounted, item["name"], note


# ══════════════════════════════════════════════════════════
# 3. ПОШУК ЦІН В ІНТЕРНЕТІ (Playwright)
# ══════════════════════════════════════════════════════════

#### НОВАЯ ВЕРСИЯ ПОИСКА В ИНТЕРНЕТЕ "как человек"
# НОВАЯ ВЕРСИЯ ДЛЯ ПОИСКА ЦЕН — ЧЕРЕЗ РЕАЛЬНЫЙ CHROME (CDP, обход капчи)
_PRICE_RE = re.compile(
    r'(?<!\d)(\d[\d\s\xa0]*(?:[.,]\d{1,2})?)\s*(?:грн\.?|UAH|₴)',
    re.IGNORECASE,
)

_PRODUCT_PRICE_SELECTORS = [
    "[data-qaid='product_price']",
    ".goods-tile__price-value",
    ".product-price__value",
    ".product__price",
    ".price-value",
    ".price__value",
    ".price--current",
    ".current-price",
    ".buy-price",
    ".product-price",
    "[data-qaid='price']",
    "[class*='ProductPrice']",
    "[class*='product-price']",
    "[class*='Price_price']",
    # широкі — як остання спроба:
    "span[class*='price']",
    "div[class*='price']",
    ".price",
]


def _median(vals: List[float]) -> Optional[float]:
    s = sorted(vals)
    n = len(s)
    if n == 0:
        return None
    mid = n // 2
    return s[mid] if n % 2 else (s[mid - 1] + s[mid]) / 2


def _parse_price_val(text: str) -> Optional[float]:
    clean = text.replace("\xa0", " ").strip()
    m = re.search(r'(\d[\d\s]*(?:[.,]\d{1,2})?)', clean)
    if not m:
        return None
    try:
        val = float(m.group(1).replace(" ", "").replace(",", "."))
        if 10 < val < 500_000:
            return val
    except ValueError:
        pass
    return None


async def _extract_price_from_product_page(page) -> Optional[float]:
    """Знімає ціну з відкритої сторінки товару."""
    for sel in _PRODUCT_PRICE_SELECTORS:
        try:
            els = await page.query_selector_all(sel)
            for el in els[:3]:
                txt = await el.text_content() or ""
                val = _parse_price_val(txt)
                if val:
                    return val
        except Exception:
            pass
    # Fallback: медіана чисел-з-грн на сторінці (не мінімум!)
    try:
        body = await page.evaluate("() => document.body.innerText")
        prices = []
        for m in _PRICE_RE.finditer(body or ""):
            raw = m.group(1).replace(" ", "").replace("\xa0", "").replace(",", ".")
            try:
                val = float(raw)
                if 10 < val < 500_000:
                    prices.append(val)
            except ValueError:
                pass
        if prices:
            return _median(prices)
    except Exception:
        pass
    return None


async def _get_google_result_urls(page, query: str, max_results: int = 5) -> List[str]:
    """Бере посилання на сайти зі сторінки результатів Google."""
    google_url = f"https://www.google.com/search?q={quote_plus(query)}&hl=uk&gl=ua&num=20"
    try:
        await page.goto(google_url, timeout=30_000, wait_until="domcontentloaded")
        await asyncio.sleep(2.5)
        await _wait_if_captcha(page)   # якщо капча — пауза, поки пройдеш вручну
        urls = await page.evaluate("""(maxResults) => {
            const seen = new Set();
            const results = [];
            for (const a of document.querySelectorAll('a[href]')) {
                const href = a.href;
                if (!href || !href.startsWith('http')) continue;
                try {
                    const host = new URL(href).hostname;
                    if (host.includes('google.') || host.includes('youtube.') ||
                        host.includes('gstatic.') || host.includes('googleapis.') ||
                        host.includes('googletagmanager') || host.includes('doubleclick')) continue;
                } catch(e) { continue; }
                if (seen.has(href)) continue;
                seen.add(href);
                results.push(href);
                if (results.length >= maxResults) break;
            }
            return results;
        }""", max_results)
        return urls or []
    except Exception as e:
        print(f"      Google error: {e}")
        return []


async def _get_prices_from_google_snippets(page) -> List[Tuple[float, str]]:
    """
    Бере ціни прямо зі сторінки результатів Google (сніпети) — головне джерело.
    Фільтрує розстрочку («/міс») і бере найбільшу реальну ціну в блоці.
    """
    try:
        results = await page.evaluate("""() => {
            const found = [];
            const priceRe = /(\\d[\\d\\s\\u00a0]*(?:[,.]\\d{1,2})?)\\s*(UAH|грн|₴)/gi;
            const junkRe = /\\/\\s*міс|\\/\\s*мес|\\bміс\\b|\\bмес\\b|на\\s*місяць|у\\s*місяць/i;
            const blocks = document.querySelectorAll('[data-hveid], .g, .tF2Cxc, [jscontroller]');
            for (const block of blocks) {
                const text = block.innerText || '';
                const link = block.querySelector('a[href]');
                if (!link) continue;
                const href = link.href;
                if (!href || href.includes('google.')) continue;
                let match;
                priceRe.lastIndex = 0;
                const candidates = [];
                while ((match = priceRe.exec(text)) !== null) {
                    const raw = match[1].replace(/[\\s\\u00a0]/g, '').replace(',', '.');
                    const val = parseFloat(raw);
                    if (!(val > 10 && val < 500000)) continue;
                    const ctx = text.slice(Math.max(0, match.index - 14),
                                           match.index + match[0].length + 10);
                    if (junkRe.test(ctx)) continue;
                    candidates.push(val);
                }
                if (candidates.length) {
                    found.push({price: Math.max.apply(null, candidates), url: href});
                }
            }
            return found;
        }""")
        return [(r["price"], r["url"]) for r in (results or [])]
    except Exception:
        return []


async def _is_captcha(page) -> bool:
    """True, якщо Google показав перевірку робота / 'unusual traffic'."""
    try:
        if "/sorry/" in (page.url or ""):
            return True
        body = (await page.evaluate("() => document.body.innerText") or "").lower()
        markers = [
            "unusual traffic", "незвичний трафік", "необычный трафик",
            "i'm not a robot", "не робот", "captcha", "recaptcha",
            "підтвердьте, що ви", "подтвердите, что вы",
        ]
        return any(m in body for m in markers)
    except Exception:
        return False


async def _wait_if_captcha(page) -> None:
    """Якщо з'явилась капча — ставимо паузу, поки людина пройде її у вікні Chrome."""
    if not await _is_captcha(page):
        return
    print(f"   {Fore.YELLOW}⛔ Google показав перевірку робота.{Style.RESET_ALL}")
    print(f"   {Fore.YELLOW}→ Перейди у вікно Chrome (порт 9222), пройди капчу,{Style.RESET_ALL}")
    print(f"   {Fore.YELLOW}  потім повернись сюди і натисни Enter...{Style.RESET_ALL}")
    try:
        input()
    except EOFError:
        await asyncio.sleep(25)
    try:
        await page.reload(timeout=30_000, wait_until="domcontentloaded")
        await asyncio.sleep(2.0)
    except Exception:
        pass


def _shop_query(item_name: str) -> str:
    """Чистий короткий запит для пошуку в магазині: прибираємо коди в дужках."""
    q = re.sub(r'\([^)]*\)', ' ', item_name)   # викидаємо артикули в дужках
    q = re.sub(r'\s+', ' ', q).strip()
    return q[:60]


async def _get_ddg_result_urls(page, query: str, max_results: int = 5) -> List[str]:
    """Запас: бере посилання на магазини з HTML-видачі DuckDuckGo (не банить ботів)."""
    ddg_url = f"https://html.duckduckgo.com/html/?q={quote_plus(query + ' ціна купити Україна')}"
    try:
        await page.goto(ddg_url, timeout=30_000, wait_until="domcontentloaded")
        await asyncio.sleep(1.5)
        urls = await page.evaluate("""(maxResults) => {
            const out = []; const seen = new Set();
            for (const a of document.querySelectorAll('a.result__a, a[href]')) {
                let href = a.href; if (!href) continue;
                try {
                    const u = new URL(href);
                    if (u.hostname.includes('duckduckgo.com')) {
                        const real = u.searchParams.get('uddg');
                        if (real) href = decodeURIComponent(real); else continue;
                    }
                } catch(e) { continue; }
                if (!href.startsWith('http')) continue;
                if (seen.has(href)) continue; seen.add(href);
                out.push(href);
                if (out.length >= maxResults) break;
            }
            return out;
        }""", max_results)
        return urls or []
    except Exception as e:
        print(f"      DuckDuckGo error: {e}")
        return []


_SERPER_KEY_CACHE = None


def get_serper_key() -> str:
    """Ключ Serper: спершу змінна оточення SERPER_API_KEY, потім файл поруч зі скриптом."""
    global _SERPER_KEY_CACHE
    if _SERPER_KEY_CACHE is not None:
        return _SERPER_KEY_CACHE
    key = os.environ.get("SERPER_API_KEY", "").strip()
    if not key:
        p = Path(__file__).with_name(SERPER_KEY_FILE)
        if p.exists():
            key = p.read_text(encoding="utf-8").strip()
    if key == "ВСТАВ_СЮДИ_КЛЮЧ_SERPER":   # незаповнений шаблон
        key = ""
    _SERPER_KEY_CACHE = key
    return key


def _serper_call(endpoint: str, query: str) -> dict:
    """Один запит до Serper (shopping/search), ринок — Україна. Помилку не валимо."""
    key = get_serper_key()
    if not key:
        return {}
    req = urllib.request.Request(
        f"https://google.serper.dev/{endpoint}",
        data=json.dumps({"q": query, "gl": "ua", "hl": "uk"}).encode("utf-8"),
        headers={"X-API-KEY": key, "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read().decode("utf-8"))
    except Exception as e:
        print(f"   {Fore.YELLOW}⚠ Serper {endpoint}: {e}{Style.RESET_ALL}")
        return {}


def _parse_uah_price(text) -> Optional[float]:
    """'1 447,18 грн' / '189,38 грн' / '₴2 333' → float. UA: пробіл=тисячі, кома=дробова."""
    if text is None:
        return None
    s = str(text).replace("\xa0", " ")
    m = re.search(r"\d[\d \.,]*", s)
    if not m:
        return None
    raw = m.group(0).strip().replace(" ", "")
    if "," in raw and "." in raw:        # напр. 1.447,18 → крапка = тисячі
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        val = float(raw)
        return val if val > 0 else None
    except ValueError:
        return None


def _first_price_in(snippet: str) -> Optional[str]:
    """Витягує перший шматок схожий на ціну зі сніпета органіки."""
    if not snippet:
        return None
    m = re.search(r"\d[\d \xa0.,]*\s*(?:грн|₴|UAH)", snippet, re.I)
    return m.group(0) if m else None


def _average_prices(prices: List[float], sources: List[str]) -> Tuple[float, str]:
    """
    Стійке усереднення. Спершу відсікаємо явні викиди (копійки-брухт і завищені
    модифікації) — лишаємо смугу навколо медіани. Потім за правилом інструкції
    відкидаємо мінімум і максимум та усереднюємо. Медіану кладемо у примітку.
    """
    prices_str = " | ".join(f"{p:,.2f}" for p in prices)
    med = _median(prices) or 0
    band = [p for p in prices if med and 0.4 * med <= p <= 2.5 * med]
    if len(band) < 2:
        band = prices[:]
    band_sorted = sorted(band)
    if len(band_sorted) >= 5:            # відкидаємо 1 мін + 1 макс (правило інструкції)
        band_sorted = band_sorted[1:-1]
    avg = round(sum(band_sorted) / len(band_sorted), 2)
    med_band = _median(band_sorted) or avg

    # розкид рахуємо по ОЧИЩЕНОМУ кластеру (без брухту), щоб флаг був осмисленим
    lo, hi = min(band_sorted), max(band_sorted)
    spread = (hi / lo) if lo > 0 else float("inf")
    flag = "⚠ ПЕРЕВІРИТИ ВРУЧНУ (великий розкид/мало даних). " if (spread > 3.0 or len(band) < 2) else ""
    note = (
        f"{flag}avg={avg:,.2f} грн | медіана={med_band:,.2f} грн "
        f"({len(band_sorted)} з {len(prices)} проп.) | всі: {prices_str} грн | "
        + " | ".join(sources)
    )
    return avg, note


def web_search_price(item_name: str) -> Optional[Tuple[float, str]]:
    """
    Ціна через Serper (API Google Shopping) — без браузера і без капчі.
    Основне джерело — /shopping (ціна готовим полем); фолбек — /search (зі сніпета).
    """
    query = re.sub(r"\s+", " ", item_name).strip()
    print(f"   {Fore.CYAN}Serper: {query[:70]}{Style.RESET_ALL}")

    found_prices: List[float] = []
    found_sources: List[str] = []

    # --- 1. Google Shopping (основне джерело) ---
    for s in _serper_call("shopping", query).get("shopping", []):
        price = _parse_uah_price(s.get("price"))
        if price is None:
            continue
        site = (s.get("source") or urlparse(s.get("link", "")).netloc).replace("www.", "")
        if any(skip in site.lower() for skip in SKIP_PRICE_SITES):
            continue
        found_prices.append(price)
        found_sources.append(f"{site}: {s.get('link', '')}")
        print(f"   {Fore.GREEN}💰 {price:,.2f} грн — {site}{Style.RESET_ALL}")

    # --- 2. Фолбек: органіка, якщо Shopping порожній ---
    if not found_prices:
        print("   ⚪ Shopping порожній — пробую органіку")
        for o in _serper_call("search", query).get("organic", []):
            price = _parse_uah_price(_first_price_in(o.get("snippet", "")))
            if price is None:
                continue
            site = urlparse(o.get("link", "")).netloc.replace("www.", "")
            if any(skip in site.lower() for skip in SKIP_PRICE_SITES):
                continue
            found_prices.append(price)
            found_sources.append(f"{site}: {o.get('link', '')}")
            print(f"   {Fore.GREEN}💰 {price:,.2f} грн — {site}{Style.RESET_ALL}")

    if not found_prices:
        return None

    return _average_prices(found_prices, found_sources)

# КОНЕЦ НОВОЙ ВЕРСИИ
# ══════════════════════════════════════════════════════════
# 4. ЧИТАННЯ / ЗАПИС GOOGLE SHEETS
# ══════════════════════════════════════════════════════════
# Эта функция проверяет заполненны ли ячейки 
def _is_empty_price(s: str) -> bool:
    t = (s or "").strip().lower()
    t = t.replace("\xa0", "").replace(" ", "").replace("грн.", "").replace("грн", "")
    t = t.replace("$", "").replace("₴", "").replace(",", ".")
    return t in ("", "0", "0.0", "0.00", "-", "—", "n/a")
# Окончена функция проверки пустых ячеек


class SpecSheet:
    """Читає вкладку «Кошторис_Наявність обладнання» і записує ціни."""

    # Ключові слова для автодетекту колонок
    _COL_KEYS = {
        "type":     ["тип товару", "тип", "категорія"],
        "name":     ["найменування", "назва", "обладнання"],
        "supplier": ["постачальник", "виробник"],
        "qty":      ["кіл-сть", "кількість", "к-сть", "к-ть"],
        "unit":     ["од. вим.", "одиниця", "од.вим"],
        "price":    ["ціна з пдв", "ціна", "price"],
        "total":    ["вартість з пдв", "вартість", "сума"],
        "notes":    ["наявність", "примітк", "коментар", "джерело", "note"],
    }

    def __init__(self, gc: gspread.Client, url: str):
        self.gc   = gc
        self.url  = url
        self.ws   = None
        self.cols: Dict[str, int] = {}   # field → 0-based column index
        self._header_row = 0             # 1-based

    # ── Відкрити таблицю ───────────────────────────────────
    def open(self) -> bool:
        try:
            wb = self.gc.open_by_url(self.url)
        except gspread.exceptions.SpreadsheetNotFound:
            print(f"{Fore.RED}❌ Таблицю не знайдено. Перевірте SHEET_URL і доступ.")
            return False
        except gspread.exceptions.APIError as ex:
            print(f"{Fore.RED}❌ Google API помилка: {ex}")
            return False

        # Знайти вкладку
        ws = None
        for sheet in wb.worksheets():
            if SPEC_TAB_NAME.lower() in sheet.title.lower():
                ws = sheet
                break

        if ws is None:
            names = [s.title for s in wb.worksheets()]
            print(f"{Fore.RED}❌ Вкладку '{SPEC_TAB_NAME}' не знайдено.")
            print(f"   Доступні вкладки: {', '.join(names)}")
            return False

        self.ws = ws
        print(f"{Fore.GREEN}✅ Відкрито вкладку: «{ws.title}»")
        return self._detect_columns()

    # ── Автодетект колонок ─────────────────────────────────
    def _detect_columns(self) -> bool:
        all_vals = self.ws.get_all_values()

        for ridx, row in enumerate(all_vals):
            low = [c.lower().strip() for c in row]
            row_text = " ".join(low)

            has_name  = any("найменування" in c or "назва" in c for c in low)
            has_price = any("ціна" in c for c in low)

            if has_name and has_price:
                self._header_row = ridx + 1
                for field, keywords in self._COL_KEYS.items():
                    for cidx, cv in enumerate(low):
                        if any(kw in cv for kw in keywords) and field not in self.cols:
                            self.cols[field] = cidx
                break

        if not self._header_row:
            print(f"{Fore.RED}❌ Не знайдено рядок заголовків у вкладці.")
            print("   Очікуються колонки: Найменування, Кіл-сть, Ціна з ПДВ")
            return False

        print(f"📌 Заголовки у рядку {self._header_row}")
        # Показуємо які колонки знайдено
        col_letters = {
            f: gspread.utils.rowcol_to_a1(1, v + 1)[:-1]
            for f, v in self.cols.items()
        }
        print(f"   Знайдені колонки: {col_letters}")

        # Перевіряємо критичні колонки
        if "price" not in self.cols:
            print(f"   {Fore.RED}❌ КРИТИЧНО: колонку 'Ціна' не знайдено — запис неможливий!")
            print(f"      Перевірте заголовки у вкладці '{SPEC_TAB_NAME}'{Style.RESET_ALL}")
            return False
        if "name" not in self.cols:
            print(f"   {Fore.RED}❌ КРИТИЧНО: колонку 'Найменування' не знайдено{Style.RESET_ALL}")
            return False

        return True

    # ── Отримати список позицій ───────────────────────────
    def get_items(self) -> List[Dict]:
        """
        Повертає позиції з кількістю та ціною.
        Пропускає розділи зі SKIP_SECTIONS та позиції зі SKIP_ITEMS.
        """
        all_vals  = self.ws.get_all_values()
        items     = []

        col_type  = self.cols.get("type",     0)
        col_name  = self.cols.get("name",     1)
        col_supp  = self.cols.get("supplier", 2)
        col_qty   = self.cols.get("qty",      3)
        col_unit  = self.cols.get("unit",     4)
        col_price = self.cols.get("price",    5)
        col_total = self.cols.get("total",    6)
        col_notes = self.cols.get("notes",    7)

        current_section = ""
        skip_section    = False

        for ridx, row in enumerate(all_vals):
            if ridx < self._header_row:          # пропускаємо шапку
                continue

            def cell(c: int) -> str:
                return str(row[c]).strip() if c < len(row) else ""

            type_val  = cell(col_type)
            name_val  = cell(col_name)
            qty_val   = cell(col_qty)
            price_val = cell(col_price)
            unit_val  = cell(col_unit)
            supp_val  = cell(col_supp)
            notes_val = cell(col_notes)

            # Рядок-розділ: немає кількості та/або назва порожня
            # if not qty_val and not name_val:
            #     # Це розділ з type_val як заголовком
            #     if type_val:
            #         current_section = type_val
            #         skip_section = any(
            #             s.lower() in type_val.lower() for s in SKIP_SECTIONS
            #         )
            #         marker = "⛔" if skip_section else "📂"
            #         print(f"\n   {marker} Розділ: {type_val}")
            #     continue

            # Рядок-підрозділ (є type_val але немає qty)
            # if type_val and not qty_val:
            #     sub_text = type_val
            #     skip_section = any(s.lower() in sub_text.lower() for s in SKIP_SECTIONS)
            #     continue

            # Порожній рядок
            if not name_val and not type_val:
                continue

            # Пропускаємо розділи інженерів
            # if skip_section:
            #     continue

            # Пропускаємо SKIP_ITEMS
            full_name = f"{type_val} {name_val}".strip()
            if any(si.lower() in full_name.lower() for si in SKIP_ITEMS):
                print(f"   ⛔ Ігнорується: {full_name[:60]}")
                continue
# ↓↓↓ Вставка Фильтра пропущенных ячее без цены↓↓↓
            if not _is_empty_price(price_val):
                print(f"   ⏭  Пропуск (ціна вже є: {price_val}): {full_name[:60]}")
                continue
            # ↑↑↑ КІНЕЦЬ ВСТАВКИ ↑↑↑

            # Кількість
            try:
                qty = float(qty_val.replace(",", ".").replace("\xa0", "")) if qty_val else 1.0
            except ValueError:
                qty = 1.0

            items.append({
                "row":           ridx + 1,        # 1-based рядок у таблиці
                "section":       current_section,
                "type":          type_val,
                "name":          name_val,
                "full_name":     full_name,
                "qty":           1.0,
                "unit":          unit_val,
                "supplier":      supp_val,
                "price_exist":   price_val,
                "notes_exist":   notes_val,
                # Індекси колонок (0-based)
                "col_price":     col_price,
                "col_supplier":  col_supp,
                "col_notes":     col_notes,
                "col_total":     col_total,
            })

        return items

    # ── Записати ціну в таблицю ───────────────────────────
    def write_result(self, item: Dict, price: float, supplier: str, note: str):
        """Пише ціну, постачальника і примітку в таблицю."""
        try:
            row = item["row"]
            updates = []

            col_price = item.get("col_price")
            col_supp  = item.get("col_supplier")
            col_notes = item.get("col_notes")

            if col_price is None:
                print(f"   {Fore.RED}❌ Не знайдено колонку 'Ціна' — пропускаю запис{Style.RESET_ALL}")
                return

            # Ціна (UAH)
            price_a1 = gspread.utils.rowcol_to_a1(row, col_price + 1)
            updates.append({"range": price_a1, "values": [[price]]})

            # Постачальник (якщо знайдено)
            if col_supp is not None:
                supp_a1 = gspread.utils.rowcol_to_a1(row, col_supp + 1)
                updates.append({"range": supp_a1, "values": [[supplier]]})

            # Примітка (якщо знайдено)
            if col_notes is not None:
                notes_a1 = gspread.utils.rowcol_to_a1(row, col_notes + 1)
                updates.append({"range": notes_a1, "values": [[note[:500]]]})

            # Batch update (один виклик API)
            self.ws.batch_update(updates, value_input_option="USER_ENTERED")
            print(f"   {Fore.GREEN}📝 Записано рядок {row}: {price:,.2f} грн → {price_a1}{Style.RESET_ALL}")

        except gspread.exceptions.APIError as ex:
            print(f"   {Fore.RED}❌ Google API помилка запису: {ex}{Style.RESET_ALL}")
            print(f"      Якщо помилка 403 — видаліть token.json і запустіть заново")
        except Exception as ex:
            print(f"   {Fore.RED}❌ Помилка запису в таблицю: {ex}{Style.RESET_ALL}")

        # Невелика пауза щоб не перевищити ліміт API
        time.sleep(0.7)


# ══════════════════════════════════════════════════════════
# 5. ГОЛОВНА ЛОГІКА
# ══════════════════════════════════════════════════════════

def _print_header():
    print(f"\n{Fore.CYAN}{'═'*62}")
    print(f"{Fore.CYAN} EScore Energy — Автоматизатор цін для СЕС")
    print(f"{Fore.CYAN}{'═'*62}{Style.RESET_ALL}\n")


async def run(sheet_url: str):
    _print_header()

    # 1. Авторизація
    creds = get_credentials()
    gc    = gspread.Client(auth=creds)

    # 2. ETI прайс
    print(f"\n{Fore.CYAN}── Крок 1: Завантаження прайсу ETI ──────────────────{Style.RESET_ALL}")
    eti = ETIPriceList(creds)
    eti_ok = eti.load()
    if not eti_ok:
        print(f"   {Fore.YELLOW}ETI пропущено, шукатимемо тільки в інтернеті{Style.RESET_ALL}")

    # Перевірка ключа Serper (інтернет-пошук цін)
    if get_serper_key():
        print(f"   {Fore.GREEN}✅ Ключ Serper знайдено — інтернет-пошук активний{Style.RESET_ALL}")
    else:
        print(f"   {Fore.RED}❌ Ключ Serper не знайдено! Встав ключ у файл '{SERPER_KEY_FILE}' "
              f"поруч зі скриптом. Інтернет-пошук цін не працюватиме.{Style.RESET_ALL}")

    # 3. Відкрити таблицю
    print(f"\n{Fore.CYAN}── Крок 2: Читання таблиці від інженерів ─────────────{Style.RESET_ALL}")
    spec = SpecSheet(gc, sheet_url)
    if not spec.open():
        return

    items = spec.get_items()
    total = len(items)

    if total == 0:
        print(f"\n{Fore.YELLOW}⚠  Позицій для обробки не знайдено.")
        print("   Перевірте назву вкладки та структуру таблиці.")
        return

    print(f"\n{Fore.CYAN}── Крок 3: Пошук цін ({total} позицій) ───────────────────{Style.RESET_ALL}")

    stats = {"eti": 0, "web": 0, "notfound": 0}

    for idx, item in enumerate(items, 1):
        name      = item["name"]
        full_name = item["full_name"]
        has_price = bool(
            item["price_exist"] and
            item["price_exist"].strip() not in ("", "0", "$0", "$0.00", "грн.0.00")
        )

        print(f"\n[{idx}/{total}] {Fore.WHITE}{full_name[:70]}{Style.RESET_ALL}")
        if item["unit"].lower() in ("уп", "упаковка"):
            print(f"   ℹ️  Одиниця: упаковка = 100 шт")

        prefix = ""
        if has_price:
            prefix = f"[Була ціна інженерів: {item['price_exist']}] "
            print(f"   {Fore.YELLOW}ℹ️  Ціна вже є ({item['price_exist']}) — перевіряю{Style.RESET_ALL}")

        # ── ETI ──────────────────────────────────────────
        # Кабельну продукцію та аксесуари НЕ шукаємо в ETI — одразу в інтернет
        skip_eti = any(k in item["type"].lower() for k in ETI_SKIP_TYPE_KEYWORDS)
        if skip_eti:
            print(f"   ⏭  Кабельна група — пропускаю ETI, шукаю в інтернеті")
        eti_result = eti.find(full_name) if (eti_ok and not skip_eti) else None

        if eti_result:
            price, matched, note = eti_result
            print(f"   {Fore.GREEN}✅ ETI: {price:,.2f} грн{Style.RESET_ALL}")
            spec.write_result(item, price, "ETI/ДС-Електро", prefix + note)
            stats["eti"] += 1
            continue

        # ── Інтернет ──────────────────────────────────────
        print(f"   🌐 Шукаю в інтернеті...")
        web_result = web_search_price(full_name)

        if web_result:
            price, note = web_result
            print(f"   {Fore.GREEN}✅ Середня ціна: {price:,.2f} грн{Style.RESET_ALL}")
            spec.write_result(item, price, "інтернет", prefix + note)
            stats["web"] += 1
        else:
            not_found = "⚠ Ціну не знайдено на перших 3 сайтах — уточнити вручну"
            print(f"   {Fore.YELLOW}⚠  Не знайдено{Style.RESET_ALL}")
            spec.write_result(item, 0, "", prefix + not_found)
            stats["notfound"] += 1

        await asyncio.sleep(3.0)   # пауза між позиціями — щоб Google рідше показував капчу

    # ── Підсумок ───────────────────────────────────────────
    print(f"\n{Fore.CYAN}{'═'*62}")
    print(f"{Fore.GREEN} ГОТОВО!")
    print(f"{Fore.WHITE}   З прайсу ETI:    {stats['eti']} поз.")
    print(f"   З інтернету:    {stats['web']} поз.")
    print(f"   Не знайдено:   {stats['notfound']} поз. (уточнити вручну)")
    print(f"{Fore.CYAN}{'═'*62}{Style.RESET_ALL}\n")


# ══════════════════════════════════════════════════════════
# ТОЧКА ВХОДУ
# ══════════════════════════════════════════════════════════

def resolve_sheet_url(sheet_url: str = "") -> str:
    """URL таблиці: аргумент → env-змінна SHEET_URL → константа SHEET_URL у файлі."""
    return (sheet_url or os.environ.get("SHEET_URL", "") or SHEET_URL or "").strip()


def run_update(sheet_url: str = "") -> None:
    """
    Єдина точка входу для оновлення цін.
    Використовується з терміналу, Streamlit-застосунку і GitHub Actions.
    Кидає ValueError при невірному URL; інакше запускає асинхронний run().
    """
    url = resolve_sheet_url(sheet_url)
    if not url:
        raise ValueError("SHEET_URL порожній: задай посилання на таблицю "
                         "(env-змінна SHEET_URL або константа у файлі).")
    if "docs.google.com/spreadsheets" not in url:
        raise ValueError("SHEET_URL не схожий на посилання Google Sheets "
                         "(очікується https://docs.google.com/spreadsheets/...).")
    asyncio.run(run(url))


def main():
    try:
        run_update()
    except ValueError as e:
        print(f"\n{Fore.RED}❌ {e}{Style.RESET_ALL}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
